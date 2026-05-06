# =============================================================================
# Riverwall Asset Register — Merged Layer Builder
# Paste into QGIS Python Console and Run.
#
# Reads:   RiverwallForValuation/Riverwalls_For_InitialRecognition.shp  (geometry)
#          RiverWalls_for_Valuation.xlsx                                 (attributes)
# Writes:  riverwall_assets.gpkg   — 37 pre-filled parent features
#          riverwall_assets.qgs    — QGIS project
#
# CRS: GDA2020 / MGA Zone 50 (EPSG:7854) — native projected CRS, metres.
#
# Child asset workflow (QField):
#   Open a parent feature form → "Child Assets" tab → tap + to digitise a
#   sub-segment. Child_Unit_ID is auto-filled with the parent's Unit_ID.
#   Alternatively, create a new feature manually and pick the parent from
#   the "Child Unit ID" dropdown.
# =============================================================================

from osgeo import ogr, osr
from qgis.core import (
    QgsProject, QgsVectorLayer,
    QgsEditorWidgetSetup, QgsDefaultValue, QgsFieldConstraints,
    QgsCoordinateReferenceSystem, QgsRelation,
)
import os

try:
    import openpyxl
except ImportError:
    raise ImportError(
        "openpyxl not found.\n"
        "In the QGIS Python Console run:\n"
        "  import subprocess, sys\n"
        "  subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'])"
    )

# ── CONFIG ────────────────────────────────────────────────────────────────────
BASE_DIR   = r"C:\Users\Damia\Documents\Projects\coastal-asset-register"
SHP_PATH   = os.path.join(BASE_DIR, "RiverwallForValuation",
                           "Riverwalls_For_InitialRecognition.shp")
XL_PATH    = os.path.join(BASE_DIR, "RiverWalls_for_Valuation.xlsx")
GPKG_PATH  = os.path.join(BASE_DIR, "riverwall_assets.gpkg")
LAYER_NAME = "Riverwall_Assets"
QGS_PATH   = os.path.join(BASE_DIR, "riverwall_assets.qgs")
# ─────────────────────────────────────────────────────────────────────────────


# ── 1. Read Excel into dict keyed by UNIT_ID ──────────────────────────────────
# Asset_Type values vary in case/spelling across sources — normalise to schema.
_AT = {
    "REVETMENT":     "Revetment",
    "WALL":          "Wall",
    "GABION":        "Gabion",
    "BEACH":         "Beach",
    "SOFT MEASURES": "Soft Measure",
    "SOFT MEASURE":  "Soft Measure",
    "BIOENGINEERING":"Bioengineering",
}

def norm_asset_type(raw):
    return _AT.get((raw or "").strip().upper(), (raw or "").strip())

wb = openpyxl.load_workbook(XL_PATH, data_only=True)
ws = wb.active
xl_headers = [c.value for c in next(ws.iter_rows(max_row=1))]

xl_data = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = dict(zip(xl_headers, row))
    uid = (rec.get("UNIT_ID") or "").strip()
    if uid:
        xl_data[uid] = rec

print(f"[1/5] Excel: {len(xl_data)} records loaded.")


# ── 2. Create output GeoPackage ───────────────────────────────────────────────
drv = ogr.GetDriverByName("GPKG")
if os.path.exists(GPKG_PATH):
    drv.DeleteDataSource(GPKG_PATH)

ds_out  = drv.CreateDataSource(GPKG_PATH)
srs_out = osr.SpatialReference()
srs_out.ImportFromEPSG(7854)  # GDA2020 / MGA Zone 50

out_lyr = ds_out.CreateLayer(LAYER_NAME, srs_out, ogr.wkbMultiLineString)

def _s(name, w=254): f = ogr.FieldDefn(name, ogr.OFTString); f.SetWidth(w); return f
def _r(name):        return ogr.FieldDefn(name, ogr.OFTReal)
def _i(name):        return ogr.FieldDefn(name, ogr.OFTInteger)
def _d(name):        return ogr.FieldDefn(name, ogr.OFTDate)

for fld in [
    _s("Unit_ID",                 100),   # NOT NULL (enforced by QGIS constraint)
    _s("Child_Unit_ID",           100),   # parent's Unit_ID — null on parent features
    _s("Unit_Description",        254),
    _s("Location_Description",    254),
    _r("Total_Length_m"),
    _s("Material",                254),   # free text — source data too varied for dropdown
    _s("Asset_Type",               50),
    _i("Construction_Year"),
    _r("Replacement_Cost_Gross"),
    _i("Condition_Rating"),
    _r("Fair_Value_Valuation"),
    _i("Useful_Life"),
    _i("Remaining_Useful_Life"),
    _r("Residual_Value"),
    _s("Depreciation_Method_Rate", 50),
    _r("Depreciation_Expense"),
    _r("Accumulated_Depreciation"),
    _d("Date_of_Valuation"),
    _s("Photo_1",                 254),
    _s("Photo_2",                 254),
    _s("Photo_3",                 254),
    _s("Inspection_Notes",        500),   # source Comments — field inspection context
]:
    out_lyr.CreateField(fld)

print("[2/5] GeoPackage schema created.")


# ── 3. Read shapefile, join Excel, write merged features ──────────────────────
ds_shp  = ogr.Open(SHP_PATH)
shp_lyr = ds_shp.GetLayer()

written   = 0
unmatched = []

for shp_feat in shp_lyr:
    uid = (shp_feat.GetField("UNIT_ID") or "").strip()
    xl  = xl_data.get(uid, {})

    if uid not in xl_data:
        unmatched.append(uid)

    geom = shp_feat.GetGeometryRef()
    if geom is None:
        continue

    # Field values — prefer Excel for text (slightly cleaner); shapefile for numeric length
    unit_desc  = (xl.get("UNIT_DESC") or shp_feat.GetField("UNIT_DESC") or "").strip()
    loc_desc   = (xl.get("LOCATION_Description") or shp_feat.GetField("LOCATION_D") or "").strip()
    material   = (xl.get("Material") or shp_feat.GetField("Material") or "").strip().rstrip("\n")
    asset_type = norm_asset_type(xl.get("Asset_Type") or shp_feat.GetField("Asset_Type") or "")
    comments   = (xl.get("Comments") or shp_feat.GetField("Comments") or "").strip().rstrip("\n")

    try:
        length_m = float(shp_feat.GetField("Length") or 0.0)
    except (TypeError, ValueError):
        length_m = 0.0

    feat_out = ogr.Feature(out_lyr.GetLayerDefn())
    feat_out.SetGeometry(ogr.ForceToMultiLineString(geom.Clone()))
    feat_out.SetField("Unit_ID",              uid)
    feat_out.SetField("Unit_Description",     unit_desc)
    feat_out.SetField("Location_Description", loc_desc)
    feat_out.SetField("Total_Length_m",       length_m)
    feat_out.SetField("Material",             material)
    feat_out.SetField("Asset_Type",           asset_type)
    feat_out.SetField("Inspection_Notes",     comments)
    # Valuation, condition, and photo fields left null — completed during assessment
    out_lyr.CreateFeature(feat_out)
    written += 1

ds_shp = None
ds_out = None  # flush and close

print(f"[3/5] {written} features written to GeoPackage.")
if unmatched:
    print(f"       WARNING — no Excel match for: {unmatched}")


# ── 4. Load layer into QGIS ───────────────────────────────────────────────────
vlyr = QgsVectorLayer(f"{GPKG_PATH}|layername={LAYER_NAME}", LAYER_NAME, "ogr")
assert vlyr.isValid(), f"Layer failed to load — check: {GPKG_PATH}"

# Field aliases (display labels in attribute form)
ALIASES = {
    "Unit_ID":                  "Unit ID *",
    "Child_Unit_ID":            "Child Unit ID (select parent)",
    "Unit_Description":         "Unit Description",
    "Location_Description":     "Location Description",
    "Total_Length_m":           "Total Length (linear m)",
    "Material":                 "Material",
    "Asset_Type":               "Asset Type",
    "Construction_Year":        "Construction Year",
    "Replacement_Cost_Gross":   "Replacement Cost – Gross Value",
    "Condition_Rating":         "Condition Rating (1–5)",
    "Fair_Value_Valuation":     "Fair Value Valuation",
    "Useful_Life":              "Useful Life",
    "Remaining_Useful_Life":    "Remaining Useful Life",
    "Residual_Value":           "Residual Value",
    "Depreciation_Method_Rate": "Depreciation Method / Rate",
    "Depreciation_Expense":     "Depreciation Expense",
    "Accumulated_Depreciation": "Accumulated Depreciation",
    "Date_of_Valuation":        "Date of Valuation",
    "Photo_1":                  "Photo 1",
    "Photo_2":                  "Photo 2",
    "Photo_3":                  "Photo 3",
    "Inspection_Notes":         "Inspection Notes (source)",
}
for fname, alias in ALIASES.items():
    idx = vlyr.fields().indexFromName(fname)
    if idx >= 0:
        vlyr.setFieldAlias(idx, alias)

# Must add to project BEFORE Value Relation widget setup (needs layer ID)
QgsProject.instance().addMapLayer(vlyr)


# ── 5. Widget configurations ──────────────────────────────────────────────────
def _idx(n):
    return vlyr.fields().indexFromName(n)

def value_map(name, values):
    cfg = {"map": [{str(v): str(v)} for v in values]}
    vlyr.setEditorWidgetSetup(_idx(name), QgsEditorWidgetSetup("ValueMap", cfg))

def date_widget(name):
    cfg = {"display_format": "yyyy-MM-dd", "field_format": "yyyy-MM-dd", "calendar_popup": True}
    vlyr.setEditorWidgetSetup(_idx(name), QgsEditorWidgetSetup("DateTime", cfg))

def attachment(name):
    cfg = {
        "DocumentViewer": 0,
        "FileWidget": True,
        "FileWidgetButton": True,
        "FileWidgetFilter": "Images (*.jpg *.jpeg *.png *.tif *.tiff)",
        "RelativeStorage": 1,   # paths relative to project file
        "StorageMode": 0,
    }
    vlyr.setEditorWidgetSetup(_idx(name), QgsEditorWidgetSetup("ExternalResource", cfg))

# Not-null constraint on Unit_ID
vlyr.setFieldConstraint(_idx("Unit_ID"), QgsFieldConstraints.ConstraintNotNull)

# Child_Unit_ID — Value Relation back to Unit_ID in this layer.
# Filter excludes features that are already children (prevents circular refs).
# Dropdown shows Unit_Description so inspectors can identify the parent by name.
vlyr.setEditorWidgetSetup(_idx("Child_Unit_ID"), QgsEditorWidgetSetup("ValueRelation", {
    "Layer":            vlyr.id(),
    "Key":              "Unit_ID",          # stored value
    "Value":            "Unit_Description", # display value
    "AllowMulti":       False,
    "AllowNull":        True,
    "FilterExpression": '"Child_Unit_ID" IS NULL OR "Child_Unit_ID" = \'\'',
    "NofColumns":       1,
    "OrderByValue":     True,
    "UseCompleter":     True,
}))

# Asset Type dropdown (normalised values)
value_map("Asset_Type", [
    "Revetment", "Wall", "Gabion", "Beach", "Soft Measure", "Bioengineering",
])

# Condition Rating 1–5
value_map("Condition_Rating", [1, 2, 3, 4, 5])

# Depreciation method
value_map("Depreciation_Method_Rate", ["Straight Line", "Diminishing Value"])

# Date of Valuation — calendar picker, default = today
date_widget("Date_of_Valuation")
vlyr.setDefaultValueDefinition(
    _idx("Date_of_Valuation"), QgsDefaultValue("to_date(now())")
)

# Photo attachment widgets
for p in ("Photo_1", "Photo_2", "Photo_3"):
    attachment(p)

print("[4/5] Widget configurations applied.")


# ── 6. Parent-child QGIS Relation (self-referencing) ─────────────────────────
# When a parent feature's form is open, the "Child Assets" tab lists all
# child features and allows new ones to be digitised directly (works in QField).
rel = QgsRelation()
rel.setId("riverwall_child_parent")
rel.setName("Child Assets")
rel.setReferencingLayer(vlyr.id())          # child: holds Child_Unit_ID
rel.addFieldPair("Child_Unit_ID", "Unit_ID")
rel.setReferencedLayer(vlyr.id())           # parent: holds Unit_ID
QgsProject.instance().relationManager().addRelation(rel)


# ── 7. Save project ───────────────────────────────────────────────────────────
QgsProject.instance().setCrs(QgsCoordinateReferenceSystem("EPSG:7854"))
QgsProject.instance().setFileName(QGS_PATH)
QgsProject.instance().write()

print("[5/5] Project saved.")
print()
print(f"  Layer   : {LAYER_NAME}  ({written} parent assets)")
print(f"  GPKG    : {GPKG_PATH}")
print(f"  Project : {QGS_PATH}")
print(f"  CRS     : GDA2020 / MGA Zone 50 (EPSG:7854)")
print()
print("QField child asset options:")
print("  A) Open a parent form → 'Child Assets' tab → tap '+' to digitise a sub-segment")
print("  B) Create a new feature → set Unit_ID (new child ID) → pick parent from 'Child Unit ID' dropdown")
